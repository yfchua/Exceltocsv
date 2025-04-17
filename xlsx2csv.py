import openpyxl
import csv
import os

def xlsx_to_csv_all_sheets(xlsx_file, output_dir):
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        csv_file = os.path.join(output_dir, f"{sheet_name}.csv")

        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"Sheet '{sheet_name}' written to {csv_file}")

# Example usage:
xlsx_to_csv_all_sheets('input.xlsx', 'output_csvs')
